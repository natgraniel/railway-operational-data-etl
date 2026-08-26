from datetime import date
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from src.extractors.pdf_extractor import PDFExtractor
from src.extractors.word_extractor import WordExtractor
from src.loaders.excel_loader import ExcelLoader, TestTrainInput
from src.models.workbook_reader import WorkbookReader
from src.transformers.transformation_layer import TransformationLayer
from src.validators.validation_layer import ValidationLayer


def test_loader_writes_multiple_test_trains_and_preserves_schedule(tmp_path: Path):
    raw = Path("data/raw")
    source = next(raw.glob("*.xlsx"))
    pdf = next(raw.glob("*.pdf"))
    docx = next(raw.glob("*.docx"))
    output = tmp_path / "Programa_actualizado.xlsx"

    pdf_result = PDFExtractor().extract(str(pdf))
    word_result = WordExtractor().extract(str(docx))
    transformed = TransformationLayer().transform(
        pdf_result.commercial_services, pdf_result.reserve, word_result.operations
    )
    # The source files vary daily. Simulate the valid PDF case where it lists
    # the physically adjacent template services in reverse order: 304-303.
    transformed.commercial_updates = [
        replace(
            update,
            registration="C004",
            source_services=("304", "303"),
        )
        if update.service in {"303", "304"}
        else update
        for update in transformed.commercial_updates
    ]
    validated = ValidationLayer().validate(transformed, WorkbookReader().read(str(source)))
    first_ticket_update = validated.ticket_updates[0]
    source_worksheet = load_workbook(source, data_only=False)["DIARIO"]
    source_schedule = source_worksheet.cell(
        row=first_ticket_update.target_row,
        column=6,
    ).value

    load_result = ExcelLoader().load(
        validated,
        source,
        output,
        program_date=date(2026, 7, 10),
        test_trains=[
            TestTrainInput("P009", "855+000 - 893+000", "R001 y R006", "17:00", "03:00"),
            TestTrainInput("P020", "893+000 - 900+000", "N001", "08:15", "10:45"),
        ],
    )
    worksheet = load_workbook(output, data_only=False)["DIARIO"]

    assert load_result.commercial_updates_written == len(
        validated.commercial_updates
    )
    assert load_result.ticket_updates_written == len(
        validated.ticket_updates
    )
    assert load_result.reserve_updates_written == len(
        validated.reserve_updates
    )
    assert load_result.test_train_written is True
    first_commercial_update = validated.commercial_updates[0]
    assert worksheet.cell(
        row=first_commercial_update.target_row,
        column=4,
    ).value == first_commercial_update.registration
    merges = {str(merged) for merged in worksheet.merged_cells.ranges}
    # The source PDF combines 301-302, whose blocks are adjacent in Programa.
    assert "D28:D34" in merges
    # The PDF may list a pair in either order; the template's physical order
    # determines the vertical merge.
    assert "D35:D38" in merges
    assert "D39:D42" in merges
    # 601-604 share an MR in the PDF but are separated by 602-603 in Programa;
    # they must remain independent to avoid covering those intervening rows.
    assert "D63:D64" in merges
    assert "D69:D70" in merges
    # The commercial merge must end before the Pruebas section header.
    assert "B75:I75" in merges
    assert "D71:D72" in merges
    assert "D71:D73" not in merges
    assert worksheet["D35"].font.name == "Noto Sans"
    assert worksheet["D35"].font.bold is True
    assert worksheet["D35"].font.sz == 12
    assert worksheet["B5"].value == "10 Julio. 2026"
    test_header_row = ExcelLoader._find_section_row(worksheet, "Pruebas")
    first_test_row = test_header_row + 1
    second_test_row = test_header_row + 2

    assert worksheet.cell(first_test_row, 2).value == "P009"
    assert worksheet.cell(first_test_row, 3).value == "855+000 - 893+000"
    assert worksheet.cell(first_test_row, 4).value == "R001 y R006"
    assert worksheet.cell(first_test_row, 5).value == "N/A"
    assert worksheet.cell(first_test_row, 6).value == "17:00"
    assert worksheet.cell(first_test_row, 7).value == "03:00"
    assert worksheet.cell(first_test_row, 8).value == "10h"

    assert worksheet.cell(second_test_row, 2).value == "P020"
    assert worksheet.cell(second_test_row, 3).value == "893+000 - 900+000"
    assert worksheet.cell(second_test_row, 4).value == "N001"
    assert worksheet.cell(second_test_row, 5).value == "N/A"
    assert worksheet.cell(second_test_row, 6).value == "08:15"
    assert worksheet.cell(second_test_row, 7).value == "10:45"
    assert worksheet.cell(second_test_row, 8).value == "2h 30m"

    assert worksheet.cell(first_test_row, 2).fill.fgColor.rgb == "00FFFFFF"
    assert worksheet.cell(second_test_row, 8).fill.fgColor.rgb == "00FFFFFF"
    assert f"H{second_test_row}:I{second_test_row}" in merges
    assert worksheet.cell(
        row=first_ticket_update.target_row,
        column=5,
    ).value == first_ticket_update.tickets_sold

    assert worksheet.cell(
        row=first_ticket_update.target_row,
        column=6,
    ).value == source_schedule
    reserve_header_row = ExcelLoader._find_section_row(worksheet, "Reserva")
    reserve_start_row = reserve_header_row + 1
    reserve_end_row = reserve_start_row + load_result.reserve_updates_written

    assert worksheet.cell(reserve_header_row, 2).value == "Reserva"

    written_registrations = [
        worksheet.cell(row, 4).value
        for row in range(reserve_start_row, reserve_end_row)
    ]
    expected_registrations = [
        update.registration
        for update in validated.reserve_updates
    ]

    assert sorted(written_registrations) == sorted(expected_registrations)

    written_statuses = [
        worksheet.cell(row, 5).value
        for row in range(reserve_start_row, reserve_end_row)
    ]
    expected_station_statuses = sum(
        " ".join(update.status.upper().split()) == "RESERVA EN ESTACION"
        for update in validated.reserve_updates
    )

    assert written_statuses.count("RESERVA EN ESTACION") == expected_station_statuses
    assert all(
        status in (None, "", "RESERVA EN ESTACION")
        for status in written_statuses
    )
    assert "RESERVA" not in written_statuses